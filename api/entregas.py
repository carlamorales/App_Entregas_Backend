import csv
import io
import os
from datetime import datetime, date

from flask import Blueprint, jsonify, request, abort
from flask_jwt_extended import get_jwt_identity
from openpyxl import load_workbook
from database import SessionLocal
from sqlalchemy import select

def create_entregas_blueprint(require_perm):
    bp = Blueprint('entregas', __name__, url_prefix='/entregas')
    ESTADOS_VALIDOS = {'PENDIENTE', 'ENTREGADO', 'CANCELADO', 'ANULADO'}
    TIPOS_CONTRATO = {'INDEFINIDO', 'PLAZO FIJO', 'CONTRATO', 'EVENTUAL'}

    def _has_data(value):
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() != ''
        return True

    def _normalize_row(row):
        normalized = {}
        for key, val in (row or {}).items():
            if key is None:
                continue
            normalized[str(key).strip().lower()] = val
        return normalized

    def _first_value(row, *keys, keep_raw=False):
        for key in keys:
            if key in row:
                val = row[key]
                if val is None:
                    continue
                if isinstance(val, str):
                    val = val.strip()
                    if val == '':
                        continue
                    return val if keep_raw else val
                if keep_raw:
                    return val
                return str(val).strip()
        return None

    def _parse_fecha(raw_fecha):
        if raw_fecha is None or (isinstance(raw_fecha, str) and raw_fecha.strip() == ''):
            return datetime.utcnow()
        if isinstance(raw_fecha, datetime):
            return raw_fecha
        if isinstance(raw_fecha, date):
            return datetime.combine(raw_fecha, datetime.min.time())
        if isinstance(raw_fecha, str):
            val = raw_fecha.strip()
            try:
                return datetime.fromisoformat(val)
            except Exception:
                for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%d/%m/%Y'):
                    try:
                        return datetime.strptime(val, fmt)
                    except Exception:
                        continue
        raise ValueError('FechaEntrega debe estar en formato ISO (YYYY-MM-DD) o ser una fecha válida de Excel')

    def _read_csv_rows(file_bytes):
        try:
            decoded = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = file_bytes.decode('latin1')
        reader = csv.DictReader(io.StringIO(decoded))
        if not reader.fieldnames:
            raise ValueError('El CSV debe incluir encabezados en la primera fila')
        rows = []
        for idx, row in enumerate(reader, start=2):  # +1 por encabezado
            normalized = _normalize_row(row)
            if not any(_has_data(v) for v in normalized.values()):
                continue
            rows.append((idx, normalized))
        return rows

    def _read_excel_rows(file_bytes):
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            raise ValueError(f'No se pudo leer el archivo Excel: {exc}') from exc
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError('El Excel está vacío')
        headers = [str(h).strip().lower() if h is not None else '' for h in header_row]
        if not any(headers):
            raise ValueError('La primera fila del Excel debe contener encabezados')

        rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for pos, header in enumerate(headers):
                if not header:
                    continue
                row_dict[header] = row[pos] if row and pos < len(row) else None
            if not any(_has_data(v) for v in row_dict.values()):
                continue
            rows.append((idx, row_dict))
        return rows

    @bp.get('/')
    @require_perm('entregas.read')
    def list_entregas():
        db = SessionLocal()
        try:
            from models import Entrega
            rows = db.execute(select(Entrega)).scalars().all()
            items = []
            for r in rows:
                items.append({
                    'ID': getattr(r, 'ID', None),
                    'Trabajador_Rut': getattr(r, 'Rut', None),
                    'Beneficio_cod': getattr(r, 'Beneficio_cod', None),
                    'Estado': getattr(r, 'Estado', None),
                })
            return jsonify(items)
        finally:
            db.close()

    

    @bp.get('/<int:id>')
    @require_perm('entregas.read')
    def get_entrega(id):
        db = SessionLocal()
        try:
            from models import Entrega
            e = db.execute(select(Entrega).where(Entrega.ID == id)).scalar_one_or_none()
            if not e:
                return abort(404)
            return jsonify({
                'ID': e.ID,
                'Trabajador_Rut': getattr(e, 'Rut', None),
                'Beneficio_cod': getattr(e, 'Beneficio_cod', None),
                'Estado': getattr(e, 'Estado', None),
                'Periodo_cod': getattr(e, 'Periodo_cod', None),
                'CodSucursal': getattr(e, 'CodSucursal', None),
                'TipoContrato': getattr(e, 'TipoContrato', None),
            })
        finally:
            db.close()

    @bp.post('/')
    @require_perm('entregas.create')
    def create_entrega():
        data = request.get_json(force=True)
        rut = data.get('Rut') or data.get('trabajador_rut') or data.get('Trabajador_Rut')
        beneficio = data.get('Beneficio_cod') or data.get('beneficio_cod')
        estado = data.get('Estado', 'PENDIENTE')
        periodo = data.get('Periodo_cod') or data.get('periodo')
        sucursal = data.get('CodSucursal') or data.get('cod_sucursal')
        tipo = data.get('TipoContrato') or data.get('tipo_contrato')
        usuario_creador = data.get('Usuario_creador')
        fecha_entrega = data.get('FechaEntrega')
        if not rut or not beneficio:
            return jsonify({'error': 'Rut y Beneficio_cod son requeridos'}), 400
        if fecha_entrega:
            try:
                fecha = datetime.fromisoformat(fecha_entrega)
            except Exception:
                return jsonify({'error': 'FechaEntrega en formato ISO requerido'}), 400
        else:
            fecha = datetime.utcnow()
        db = SessionLocal()
        try:
            from models import Entrega
            e = Entrega(Rut=rut, FechaEntrega=fecha, Beneficio_cod=beneficio, Estado=estado, Periodo_cod=periodo, CodSucursal=sucursal, TipoContrato=tipo, Usuario_creador=usuario_creador)
            db.add(e)
            db.commit()
            return jsonify({'ok': True, 'ID': e.ID}), 201
        except Exception as ex:
            db.rollback()
            return jsonify({'error': str(ex)}), 500
        finally:
            db.close()

    @bp.post('/importar')
    @require_perm('entregas.create')
    def importar_entregas():
        upload = request.files.get('file')
        if not upload or not upload.filename:
            return jsonify({'error': 'Envía el archivo en multipart/form-data con la clave "file" (.csv o .xlsx)'}), 400

        ext = os.path.splitext(upload.filename)[1].lower()
        if ext not in ('.csv', '.xlsx'):
            return jsonify({'error': 'Formato no soportado. Usa .csv o .xlsx'}), 400

        file_bytes = upload.read()
        if not file_bytes:
            return jsonify({'error': 'El archivo está vacío'}), 400

        try:
            raw_rows = _read_csv_rows(file_bytes) if ext == '.csv' else _read_excel_rows(file_bytes)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'error': f'No se pudo procesar el archivo: {exc}'}), 400

        parsed_rows = []
        errores = []
        for row_num, row in raw_rows:
            rut = _first_value(row, 'rut', 'trabajador_rut', 'rut_trabajador')
            beneficio = _first_value(row, 'beneficio_cod', 'beneficio', 'cod_beneficio')
            periodo = _first_value(row, 'periodo', 'periodo_cod', 'cod_periodo')
            sucursal = _first_value(row, 'cod_sucursal', 'sucursal')
            tipo = _first_value(row, 'tipo_contrato', 'tipocontrato')
            estado = _first_value(row, 'estado')
            usuario_creador = _first_value(row, 'usuario_creador', 'usuario_id', 'user_id') or get_jwt_identity()
            fecha_raw = _first_value(row, 'fecha_entrega', 'fecha', 'fechaentrega', keep_raw=True)

            if not rut or not beneficio:
                errores.append({'fila': row_num, 'error': 'rut y beneficio_cod son requeridos'})
                continue

            try:
                fecha_entrega = _parse_fecha(fecha_raw)
            except ValueError as exc:
                errores.append({'fila': row_num, 'error': str(exc)})
                continue

            estado_final = (estado or 'PENDIENTE').upper()
            if estado_final not in ESTADOS_VALIDOS:
                errores.append({'fila': row_num, 'error': f"Estado inválido: {estado_final}"})
                continue

            tipo_final = None
            if tipo:
                tipo_final = str(tipo).strip().upper()
                if tipo_final not in TIPOS_CONTRATO:
                    errores.append({'fila': row_num, 'error': f"TipoContrato inválido: {tipo_final}"})
                    continue

            usuario_final = None
            if usuario_creador is not None and str(usuario_creador).strip() != '':
                try:
                    usuario_final = int(usuario_creador)
                except Exception:
                    errores.append({'fila': row_num, 'error': 'usuario_creador debe ser numérico'})
                    continue

            parsed_rows.append({
                'fila': row_num,
                'rut': rut,
                'beneficio': beneficio,
                'periodo': periodo,
                'sucursal': sucursal,
                'tipo': tipo_final,
                'estado': estado_final,
                'fecha': fecha_entrega,
                'usuario_creador': usuario_final
            })

        if not parsed_rows:
            return jsonify({'error': 'No hay filas válidas en el archivo', 'errores': errores}), 400

        db = SessionLocal()
        created = []
        try:
            from models import Entrega
            for item in parsed_rows:
                try:
                    ent = Entrega(
                        Rut=item['rut'],
                        Beneficio_cod=item['beneficio'],
                        FechaEntrega=item['fecha'],
                        Estado=item['estado'],
                        Periodo_cod=item['periodo'],
                        CodSucursal=item['sucursal'],
                        TipoContrato=item['tipo'],
                        Usuario_creador=item['usuario_creador'],
                    )
                    db.add(ent)
                    db.commit()
                    created.append({'fila': item['fila'], 'id': ent.ID})
                except Exception as row_exc:
                    db.rollback()
                    errores.append({'fila': item['fila'], 'error': str(row_exc)})

            status = 200 if created else 400
            return jsonify({
                'ok': bool(created),
                'creadas': len(created),
                'insertadas': created,
                'errores': errores
            }), status
        finally:
            db.close()

    @bp.put('/<int:id>')
    @require_perm('entregas.update')
    def update_entrega(id):
        data = request.get_json(force=True)
        db = SessionLocal()
        try:
            from models import Entrega
            e = db.execute(select(Entrega).where(Entrega.ID == id)).scalar_one_or_none()
            if not e:
                return abort(404)
            for field, attr in (('Estado','Estado'), ('Periodo_cod','Periodo_cod'), ('CodSucursal','CodSucursal'), ('TipoContrato','TipoContrato')):
                if field in data:
                    setattr(e, attr, data.get(field))
            db.commit()
            return jsonify({'ok': True, 'ID': e.ID}), 200
        except Exception as ex:
            db.rollback()
            return jsonify({'error': str(ex)}), 500
        finally:
            db.close()

    return bp
